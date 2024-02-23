import os
import glob
import shutil
import subprocess
from openpyxl import *
from copy import copy
del open
path = os.getcwd()
combining_path = path + "/combining/"
os.chdir(combining_path)
from combining import combining
from combining import catcher
os.chdir(path)
import recode_finder as rf
from openpyxl.styles import Font


tmp = os.getcwd()
if os.name == "nt":
    sep = "\\"
    pre = "python3"
else:
    sep = "/"
    pre = "py"
input_path = path + "/INPUT/"
output_path = path + "/OUTPUT/"
combined_file = glob.glob(os.path.join(input_path, "*.xlsx"))[0].split(sep)[-1]
os.chdir(output_path)

wb = load_workbook("reconciling.xlsx")
wb2 = load_workbook(combined_file)

#Beginning of Coleman Function
print("\n\nWould you like to reconcile simple trials (trials with only onset/offset number errors) automatically? y/n \n")
autoreconcile = input()

if autoreconcile == "y" or autoreconcile == "Y":
    for sheet in wb:
        trial = int(sheet.title.split()[1])
        #look letters and on/offsets storage. Makes an array for each column of the reconciling.xlsx spreadsheet and populates it for the current trial
        coder1letters=[]
        coder1ons=[]
        coder1offs=[]
        coder2letters=[]
        coder2ons=[]
        coder2offs=[]
        coder3letters=[]
        coder3ons=[]
        coder3offs=[]
        for i in range(0, 3):  #iterates through coder 1 2 and 3 of that trial, collecting the coders' data in groups (column ABC, column EFG, column IJK)
            if i == 0:
                col = "E"
                cols = ["E", "F", "G"]
            elif i == 1:
                col = "I"
                cols = ["I", "J", "K"]
            else:
                col = "A"
                cols = ["A", "B", "C"]
            for c in cols:
                if c == "G" or c=="K" or c=="C": #accounts for the offset column skipping the first row (B look)
                    index = 3
                else:
                    index = 2
                while sheet[c+str(index)].value: #while there are still values in that column of the coder sheet, collects letters and on/off sets for each look
                    if i==0:
                        if c == "E":
                            coder1letters.append(sheet[c+str(index)].value)
                        elif c == "F":
                            coder1ons.append(sheet[c+str(index)].value)
                        elif c == "G":
                            coder1offs.append(sheet[c+str(index)].value)
                    elif i==1:
                        if c=="I":
                            coder2letters.append(sheet[c+str(index)].value)
                        elif c=="J":
                            coder2ons.append(sheet[c+str(index)].value)
                        elif c=="K":
                            coder2offs.append(sheet[c+str(index)].value)
                    else:
                        if c=="A":
                            coder3letters.append(sheet[c+str(index)].value)
                        elif c=="B":
                            coder3ons.append(sheet[c+str(index)].value)
                        elif c=="C":
                            coder3offs.append(sheet[c+str(index)].value)
                    index += 1

#Checks for trials with either 1&3 in agreement or 2&3 in agreement on look type/order, adds the looks to the incorrect coder (only when coders in agreement have more looks than incorrect coder)   
        if coder1letters==coder3letters and coder1letters!=coder2letters and len(coder1letters) >= len(coder2letters): #if coder1 and 3 have look agreement, 1 and 2 don't, and 1 has more looks than 2
            for i in range(0,len(coder1letters)):
                if coder1letters[i]!=coder2letters[i] and len(coder1letters)>len(coder2letters): #if 1 and 2 don't agree on this look (index i), and 1 still has more looks than 2, then we insert a look into coder 2
                    print('Added a ' + coder3letters[i] + ' look with onset ' + str(coder3ons[i]) + ' in ' + str(sheet.title))
                    sheet.move_range(cell_range='I' + str(i+2) + ':K99', rows = 1, cols = 0) #shifts the looks below the disagreeing one down a row to make room for the insertion
                    for j in range(0,3): #goes column by column, inserting the letter, onset, and offset from coder 3, which are then turned greent by Font
                        if j == 0:
                            sheet['I'+str(i+2)].value = sheet['A'+str(i+2)].value
                            sheet['I'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        elif j ==1:
                            sheet['J'+str(i+2)].value = sheet['B'+str(i+2)].value
                            sheet['J'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        else:
                            sheet['K'+str(i+2)].value = sheet['C'+str(i+2)].value
                            sheet['K'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                    if sheet['J'+str(i+3)].value < sheet['K'+str(i+2)].value: #If the onset right after that offset comes before the replacement, text is made blue (explained more in documentation)
                        sheet['K'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF")
                    
                    coder2letters.insert(i, coder3letters[i]) #now that spreadsheet has been changed, adjustment is made to the arrays storing these values so that they reflect the sheet
                    coder2ons.insert(i, coder3ons[i])
                    coder2offs.insert(i-1, coder3offs[i-1])
                    coder3letters[i] = coder3letters[i]+"X" #look is marked with an X (in the array only) so that the reconciler will not attempt to reuse it
                    wb.save("reconciling.xlsx")
                elif coder1letters[i]!=coder2letters[i]: #if 1 and 2 don't agree on look order, and 1 and 2 are equal, we replace that look with coder 3 rather than inserting one
                    print('Replaced with a ' + coder3letters[i] + ' look with onset ' + str(coder3ons[i]) + ' in trial ' + str(sheet.title))
                    for j in range(0,3): #goes column by column, replacing with the letter, onset, and offset from coder 3, which are then turned greent by Font
                        if j == 0:
                            sheet['I'+str(i+2)].value = sheet['A'+str(i+2)].value #replace letter 
                            sheet['I'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        elif j ==1:
                            sheet['J'+str(i+2)].value = sheet['B'+str(i+2)].value #replace onset
                            sheet['J'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        else:
                            sheet['K'+str(i+2)].value = sheet['C'+str(i+2)].value #replace offset
                            sheet['K'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                    if sheet['J'+str(i+3)].value < sheet['K'+str(i+2)].value:  #If the onset right after that offset comes before the replacement, text is made blue (explained more in documentation)
                        sheet['K'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF")
                        
                    #now that spreadsheet has been changed, adjustment is made to the arrays storing these values so that they reflect the sheet    
                    coder2letters[i] = coder3letters[i] 
                    coder2ons[i]= coder3ons[i]
                    coder2offs[i-1] = coder3offs[i-1]
                    coder3letters[i] = coder3letters[i]+"X" #look is marked with an X (in the array only) so that the reconciler will not attempt to reuse it
                    wb.save("reconciling.xlsx")



       #performs the same exact function as above, except for errors in Coder 1 rather than in Coder 2             
        elif coder2letters==coder3letters and coder2letters!=coder1letters and len(coder2letters)>=len(coder1letters):
            for i in range(0,len(coder2letters)):
                #print(coder1letters[i] + coder2letters[i])
                if coder1letters[i]!=coder2letters[i] and len(coder2letters)>len(coder1letters):
                    print('Added a ' + coder3letters[i] + ' look with onset ' + str(coder3ons[i]) + ' in ' + str(sheet.title))
                    sheet.move_range(cell_range='E' + str(i+2) + ':G99', rows = 1, cols = 0)
                    for j in range(0,3):
                        if j == 0:
                            sheet['E'+str(i+2)].value = sheet['A'+str(i+2)].value
                            sheet['E'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        elif j ==1:
                            sheet['F'+str(i+2)].value = sheet['B'+str(i+2)].value
                            sheet['F'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        else:
                            sheet['G'+str(i+2)].value = sheet['C'+str(i+2)].value
                            sheet['G'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                    if sheet['F'+str(i+3)].value < sheet['G'+str(i+2)].value:
                        sheet['G'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF")
                        
                    coder1letters.insert(i, coder3letters[i])
                    coder1ons.insert(i, coder3ons[i])
                    coder1offs.insert(i-1,coder3offs[i-1])
                    coder3letters[i] = coder3letters[i]+"X"
                    wb.save("reconciling.xlsx")

                elif coder1letters[i]!=coder2letters[i]:
                    print('Replaced with a ' + coder3letters[i] + ' look with onset ' + str(coder3ons[i]) + ' in trial ' + str(sheet.title))
                    
                    for j in range(0,3):
                        if j == 0:
                            sheet['E'+str(i+2)].value = sheet['A'+str(i+2)].value
                            sheet['E'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        elif j ==1:
                            sheet['F'+str(i+2)].value = sheet['B'+str(i+2)].value
                            sheet['F'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        else:
                            sheet['G'+str(i+2)].value = sheet['C'+str(i+2)].value
                            sheet['G'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                    if sheet['F'+str(i+3)].value < sheet['G'+str(i+2)].value:
                        sheet['G'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF")
                        
                    coder1letters[i] =coder3letters[i]
                    coder1ons[i] = coder3ons[i]
                    coder1offs[i-1] = coder3offs[i-1]
                    coder3letters[i] = coder3letters[i]+"X"
                    wb.save("reconciling.xlsx")
                
                    
#makes changes if the order of looks is correct for both coders, but onsets/offsets need fixed.  threshold for change is x>3 frames
        if coder1letters==coder2letters and coder1letters==coder3letters: #if all three coders have look agreement (agree on type/order)
            #check onsets
            for i in range(0,len(coder1ons)): 
                if(abs(coder1ons[i]-coder2ons[i])>3) and ("X" not in coder3letters[i]): #if coder 1&2 disagree by more than 3 frames and that look hasn't already been used for a full look replacement above
                    if(abs(coder1ons[i]-coder3ons[i])<=3): #if coder 3 is within 3 frames of coder 1
                        #change 2 into 3
                        print('Adjusted Coder 2 for onset ' + str(i+1) + ' in ' + str(sheet.title))#let the user know what was changed
                        sheet['J'+str(i+2)].value = coder3ons[i] #replaces value in green
                        sheet['J'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878")
                        if sheet['K'+str(i+2)].value<sheet['J'+str(i+2)].value: # if the following offset now comes before its onset chronologically (has a lower frame number) the font is changed to blue
                            sheet['J'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF")
                    elif(abs(coder2ons[i]-coder3ons[i])<=3):
                        #change 1 into 3
                        print('Adjusted Coder 1 for onset ' + str(i+1) + ' in ' + str(sheet.title))#let the user know what was changed
                        sheet['F'+str(i+2)].value = coder3ons[i] #replaces value in green
                        sheet['F'+str(i+2)].font = Font(name="Calibri", size=11, color="50C878")
                        if sheet['G'+str(i+2)].value<sheet['F'+str(i+2)].value: # if the following offset now comes before its onset chronologically (has a lower frame number) the font is changed to blue
                            sheet['F'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF")
            #check offsets
            for i in range(0,len(coder1offs)): #for all coder 1 offsets
                if(abs(coder1offs[i]-coder2offs[i])>3):  #if coder 1 and 2 disagree by more than 3 frames
                    if(abs(coder1offs[i]-coder3offs[i])<=3): #ifcoder 3 is within 3 frames of coder 1 (both agree)
                        #change 2 into 3
                        #if ("X" not in coder3letters[i]):
                        print('Adjusted Coder 2 for offset ' + str(i+1) + ' in ' + str(sheet.title)) #let the user know what was changed
                        sheet['K'+str(i+3)].value = coder3offs[i] #set coder 3 value to coder 2
                        sheet['K'+str(i+3)].font = Font(name="Calibri", size=11, color = "50C878") #make it green
                        if sheet['J'+str(i+4)].value<sheet['K'+str(i+3)].value: # if the order is off chronologically
                            sheet['K'+str(i+3)].font = Font(name="Calibri", size=11, color = "0000FF") #make it blue
                    elif(abs(coder2offs[i]-coder3offs[i])<=3): #if coder 3 is within 3 frames of coder 2 (both agree)
                        #change 1 into 3
                        if ("X" not in coder3letters[i]):
                            print('Adjusted Coder 1 for offset ' + str(i+1) + ' in ' + str(sheet.title))#let the user know what was changed
                            sheet['G'+str(i+3)].value = coder3offs[i] #set coder 3 value to coder 1
                            sheet['G'+str(i+3)].font = Font(name="Calibri", size=11, color = "50C878") #make it green
                        if sheet['F'+str(i+4)].value<sheet['G'+str(i+3)].value: #if order is off chronologically
                            sheet['G'+str(i+3)].font = Font(name="Calibri", size=11, color = "0000FF") #make it blue

#checks for trials with 1&2 in agreement on look type/order but not 3, pulls on/offsets from 3 to correct 1 or 2
        if coder1letters==coder2letters and coder1letters!=coder3letters: #if 1/2 have look agreement but 3 does not
            for i in range(0,len(coder1ons)): #for all looks (length of onsets)
                if(abs(coder1ons[i]-coder2ons[i])>3) and ("X" not in coder3letters[i]): #if coder 1 and 2 disagree by >3 frames and the look wasn't used for a look insertion above
                    for j in range(0,len(coder3ons)): #for all onsets
                        if coder3letters[j] == coder1letters[i] and abs(coder3ons[j]-coder1ons[i])<=3: #if 1/3 have look agreement and onset agreement
                            #change 2 to 3
                            print('Adjusted Coder 2 for onset ' + str(i+1) + ' in ' + str(sheet.title))#let the user know what was changed
                            sheet['J'+str(i+2)].value = coder3ons[j] #make 2 onset 3's onset
                            sheet['J'+str(i+2)].font = Font(name="Calibri", size=11, color = "50C878") #make it green
                            if coder2ons[i+2]<sheet['J'+str(i+2)].value: #if order is off chronoligcally 
                                sheet['J'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF") #make it blue
                        elif coder3letters[j] == coder2letters[i] and abs(coder3ons[j]-coder2ons[i])<=3: # if 2/3 have look agreement and onset agreement
                            #change 1 to 3
                            print('Adjusted Coder 1 for onset ' + str(i+1) + ' in ' + str(sheet.title))#let the user know what was changed
                            sheet['F'+str(i+2)].value = coder3ons[j]    #set 1 value to 3 value for onset
                            sheet['F'+str(i+2)].font = Font(name="Calibri", size=11, color="50C878") #make it green
                            if coder1ons[i+2]<sheet['F'+str(i+2)].value:    #if off chronologically 
                                sheet['F'+str(i+2)].font = Font(name="Calibri", size=11, color = "0000FF") #make it blue
            for i in range(0,len(coder1offs)): #for all onsets
                if abs(coder1offs[i]-coder2offs[i])>3: #if 1/2 disagree by more than 3 frames
                    for j in range(0,len(coder3offs)): #for all ofsets
                        if (coder3letters[j+1] == coder1letters[i+1]) and (abs(coder3offs[j]-coder1offs[i])<=3): #if 1/3 have look agreement and offset agreement
                            #change 2 to 3
                            if ("X" not in coder3letters[i]):   #if this look wasn't used for a full look replacement above
                                print('Adjusted Coder 2 for offset ' + str(i+1) + ' in ' + str(sheet.title))#let the user know what was changed
                                sheet['K'+str(i+3)].value = coder3offs[j]   #replace 2 offset with 3's offset
                                sheet['K'+str(i+3)].font = Font(name="Calibri", size=11, color = "50C878") #make it green
                            if sheet['J'+str(i+4)].value <sheet['K'+str(i+3)].value: # if order is off chronologically 
                                sheet['K'+str(i+3)].font = Font(name="Calibri", size=11, color = "0000FF") #make it blue
                        elif (coder3letters[j+1] == coder2letters[i+1]) and (abs(coder3offs[j]-coder2offs[i])<=3): #if 3/2 have look agreement and offset agreement
                            #change 1 to 3
                            if ("X" not in coder3letters[i]): #if the look wasn't used for a full look replacement above
                                print('Adjusted Coder 1 for offset ' + str(i+1) + ' in ' + str(sheet.title))#let the user know what was changed
                                sheet['G'+str(i+3)].value = coder3offs[j]   #change spreadsheet from 1 offset to 3's offset
                                sheet['G'+str(i+3)].font = Font(name="Calibri", size=11, color = "50C878") #make it green
                            if sheet['F'+str(i+4)].value <sheet['G'+str(i+3)].value: #if order is off chronologically 
                                sheet['G'+str(i+3)].font = Font(name="Calibri", size=11, color = "0000FF") #make it blue



        wb.save("reconciling.xlsx") #save all changes
                

#End of Coleman's Function


            
for i in range(0, 2):  #sets i equal to either 0 or 1, and iterates through either coder 1 or coder 2 in reconciling.xslx
    if i == 0:
        col = "E"
        cols = ["E", "F", "G"]
    else:
        col = "I"
        cols = ["I", "J", "K"]

    # finds the B's and S's in the combined file
    sheet2 = wb2.worksheets[i]
    coder = sheet2.title
    locations = {}
    trial_num = 1
    for j in range(1, sheet2.max_row+1):#iterate through entire sheet in combined sheet
        cell = sheet2["A" + str(j)]
        if cell.value == "B":
            locations[trial_num] = {}
            locations[trial_num]["B"] = j  #stores location information for start of trial in set locations
        if cell.value == "S":
            try:
                locations[trial_num]["S"] = j
            except:
                print("ERROR: There's probably missing B in " + coder)
                exit(1)
            trial_num += 1
            
    # handles the reconciling file
    for sheet in wb:   # for sheet in reconciling.xslx
        trial = int(sheet.title.split()[1]) #assigns trial number from sheet title
        coder = sheet[col+"1"].value #pulls coder number from previously assigned columns
        b_loc = locations[trial]["B"]
        s_loc = locations[trial]["S"]
        old_length = s_loc - b_loc + 1 #subtracting row numbers of B and S found in combined file 
        index = 1 #starts at header
        while sheet[col+str(index)].value: #while there are still values in that column of the coder sheet (E or I for 1 or 2 respectively)
            index += 1
        new_length = index-2 #number of rows in coder's trial
        diff = new_length - old_length
        if diff < 0:
            wb2[coder].delete_rows(b_loc, -diff)
        if diff > 0:
            wb2[coder].insert_rows(b_loc, diff)
        for key in locations:
            if key > trial:
                locations[key]["B"] += diff
                locations[key]["S"] += diff
        index = 2
        for j in range(b_loc, b_loc + new_length):
            for k in [["A", 0], ["B", 1], ["C", 2]]:
                wb2[coder][k[0]+str(j)].value = sheet[cols[k[1]]+str(index)].value
                wb2[coder][k[0]+str(j)].fill = copy(sheet[cols[k[1]]+str(index)].fill)
                wb2[coder][k[0]+str(j)].font = copy(sheet[cols[k[1]]+str(index)].font)
            index += 1
        
    wb2.save(combined_file)
        
print("Data added successfully")

# run the combining
shutil.copyfile(combined_file, combining_path + "/input/" + combined_file)
os.chdir(combining_path)
catcher.main
combining.main()
os.chdir(combining_path + "/input/")
shutil.copyfile(combined_file, output_path + combined_file)

# check for bad trials
path = tmp
os.chdir(path)
rf.main(combined_file)
